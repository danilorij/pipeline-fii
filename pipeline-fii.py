#!/usr/bin/env python3
# -*- coding: utf-8 -*-
#python -m pip install selenium beautifulsoup4 pandas numpy webdriver-manager openpyxl
"""Pipeline de processamento de dados de Fundos Imobiliários - Versão Excel"""

import os
import time
from io import StringIO
import numpy as np
import pandas as pd
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment
from openpyxl import Workbook
import openpyxl

# ====================== CONFIGURAÇÕES ======================
# Configuração do caminho de salvamento
USER_DOCUMENTS = os.path.join(os.path.expanduser('~'), 'Documents')
OUTPUT_FOLDER = os.path.join(USER_DOCUMENTS, 'Dados_FII')
os.makedirs(OUTPUT_FOLDER, exist_ok=True)
OUTPUT_FILE = os.path.join(OUTPUT_FOLDER, 'fiis_filtrados.xlsx')

# Configurações do navegador
CHROME_OPTIONS = [
    "--headless",
    "--disable-dev-shm-usage",
    "--no-sandbox",
    "--window-size=1920,1080"
]
WAIT_TIME = 5
URL_TARGET = "https://www.fundsexplorer.com.br/ranking"

# Colunas para selecionar
COLUMNS_TO_SELECT = [
    'Fundos', 'Setor', 'Preço Atual (R$)', 'Liquidez Diária (R$)',
    'Último Dividendo', 'Dividend Yield', 'DY (3M) Acumulado',
    'DY (6M) Acumulado', 'DY (12M) Acumulado', 'DY (3M) média',
    'DY (6M) média', 'DY (12M) média', 'DY Ano', 'Variação Preço',
    'Rentab. Acumulada', 'Patrimônio Líquido', 'P/VP',
    'Quant. Ativos', 'Num. Cotistas'
]

# Filtros para aplicar
FILTER_CONDITIONS = {
    'P/VP': (0.80, None),
    'Num. Cotistas': (15000, None),
    'Patrimônio Líquido': (200000000, None),
    'Liquidez Diária (R$)': (400000, None),
    'Dividend Yield': (0.65, None)
}

# ====================== FUNÇÕES AUXILIARES ======================
def configure_selenium_driver() -> webdriver.Chrome:
    """Configura e retorna o driver do Selenium"""
    options = Options()
    for option in CHROME_OPTIONS:
        options.add_argument(option)
    
    service = Service(ChromeDriverManager().install())
    return webdriver.Chrome(service=service, options=options)

def remove_popups(driver: webdriver.Chrome) -> None:
    """Remove elementos que sobrepõem o conteúdo"""
    removal_script = """
        const elementsToRemove = [
            "iframe[aria-label='Popup CTA']",
            "div[id^='hs-overlay-cta-']",
            "div[id^='hs-interactives-modal-overlay']"
        ];
        
        elementsToRemove.forEach(selector => {
            const element = document.querySelector(selector);
            if (element) element.remove();
        });
    """
    driver.execute_script(removal_script)

def select_all_columns(driver: webdriver.Chrome) -> None:
    """Seleciona todas as colunas na tabela de classificação"""
    driver.execute_script("""
        const selectButton = document.querySelector("#colunas-ranking__select-button");
        if (selectButton) selectButton.click();
    """)
    time.sleep(2)
    
    driver.execute_script("""
        const selectAll = document.querySelector("label[for='colunas-ranking__todos'] span.checkmark");
        if (selectAll) selectAll.click();
    """)
    time.sleep(2)

def scrape_fii_table() -> pd.DataFrame:
    """Extrai a tabela de dados de FIIs e retorna um DataFrame"""
    driver = configure_selenium_driver()
    try:
        driver.get(URL_TARGET)
        time.sleep(WAIT_TIME)
        
        remove_popups(driver)
        select_all_columns(driver)
        
        html_content = driver.page_source
        soup = BeautifulSoup(html_content, "html.parser")
        table = soup.find("table")
        
        return pd.read_html(StringIO(str(table)))[0]
    finally:
        driver.quit()

def normalize_column(df: pd.DataFrame, column_name: str, divisor: float = 1) -> pd.DataFrame:
    """Normaliza colunas numéricas removendo caracteres especiais"""
    if column_name not in df.columns:
        return df
    
    df[column_name] = (
        df[column_name]
        .astype(str)
        .str.replace('%', '', regex=False)
        .str.replace('.', '', regex=False)
        .str.replace(',', '.', regex=False)
    )
    
    df[column_name] = pd.to_numeric(df[column_name], errors='coerce')
    return df.assign(**{column_name: df[column_name] / divisor})

def clean_fii_data(raw_df: pd.DataFrame) -> pd.DataFrame:
    """Realiza limpeza e transformação nos dados de FIIs"""
    # Normalização de colunas
    conversion_params = [
        ('Preço Atual (R$)', 100),
        ('Último Dividendo', 100),
        ('Volatilidade', 100),
        ('P/VP', 1000),
        ('P/VPA', 100)
    ]
    
    for col, divisor in conversion_params:
        raw_df = normalize_column(raw_df, col, divisor)
    
    # Filtro de colunas
    filtered_df = raw_df[COLUMNS_TO_SELECT].copy()
    
    # Conversão de tipos
    percentage_cols = [
        'Dividend Yield', 'DY (3M) Acumulado', 'DY (6M) Acumulado',
        'DY (12M) Acumulado', 'DY (3M) média', 'DY (6M) média',
        'DY (12M) média', 'DY Ano', 'Variação Preço', 'Rentab. Acumulada'
    ]
    
    for col in percentage_cols:
        if col in filtered_df.columns:
            filtered_df[col] = (
                filtered_df[col]
                .astype(str)
                .str.replace('%', '', regex=False)
                .str.replace(',', '.', regex=False)
                .pipe(pd.to_numeric, errors='coerce')
            )
    
    numeric_cols = ['Liquidez Diária (R$)', 'Patrimônio Líquido']
    for col in numeric_cols:
        if col in filtered_df.columns:
            filtered_df[col] = (
                filtered_df[col]
                .astype(str)
                .str.replace('.', '', regex=False)
                .str.replace(',', '.', regex=False)
                .pipe(pd.to_numeric, errors='coerce')
            )
    
    if 'Num. Cotistas' in filtered_df.columns:
        filtered_df['Num. Cotistas'] = (
            filtered_df['Num. Cotistas']
            .astype(str)
            .str.replace('.', '', regex=False)
            .pipe(pd.to_numeric, errors='coerce')
        )
    
    # Definição de tipos
    dtype_mapping = {
        'Fundos': 'string',
        'Setor': 'category',
        'Preço Atual (R$)': np.float64,
        'Liquidez Diária (R$)': np.float64,
        'Último Dividendo': np.float64,
        'Dividend Yield': np.float64,
        'DY (3M) Acumulado': np.float64,
        'DY (6M) Acumulado': np.float64,
        'DY (12M) Acumulado': np.float64,
        'DY (3M) média': np.float64,
        'DY (6M) média': np.float64,
        'DY (12M) média': np.float64,
        'DY Ano': np.float64,
        'Variação Preço': np.float64,
        'Rentab. Acumulada': np.float64,
        'Patrimônio Líquido': np.float64,
        'P/VP': np.float64,
        'Quant. Ativos': 'Int64',
        'Num. Cotistas': 'Int64'
    }
    
    for col, dtype in dtype_mapping.items():
        if col in filtered_df.columns:
            filtered_df[col] = filtered_df[col].astype(dtype, errors='ignore')
    
    return filtered_df

def apply_filters(df: pd.DataFrame) -> pd.DataFrame:
    """Aplica filtros ao DataFrame conforme condições pré-definidas"""
    for column, (min_val, max_val) in FILTER_CONDITIONS.items():
        if column not in df.columns:
            continue
            
        if min_val is not None:
            df = df[df[column] >= min_val]
        if max_val is not None:
            df = df[df[column] <= max_val]
    
    return df

def export_to_excel(df: pd.DataFrame) -> None:
    """Exporta DataFrame para arquivo Excel com formatação profissional"""
    # Cria um novo workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Fundos Imobiliários"
    
    # Adiciona os dados
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    # Formatação
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = openpyxl.styles.PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    alignment = Alignment(horizontal="center", vertical="center")
    
    # Formata cabeçalho
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment
    
    # Ajusta largura das colunas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Formata números
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
    
    # Salva o arquivo
    wb.save(OUTPUT_FILE)
    print(f"Planilha Excel salva em: {OUTPUT_FILE}")

# ====================== FLUXO PRINCIPAL ======================
def main() -> None:
    """Fluxo principal de execução do pipeline"""
    print("Iniciando coleta de dados de Fundos Imobiliários...")
    
    try:
        # Coleta de dados
        print("> Extraindo dados do site...")
        raw_data = scrape_fii_table()
        
        # Processamento de dados
        print("> Processando e limpando dados...")
        cleaned_data = clean_fii_data(raw_data)
        
        # Filtragem
        print("> Aplicando filtros...")
        filtered_data = apply_filters(cleaned_data)
        
        # Exportação
        print("> Exportando para Excel...")
        export_to_excel(filtered_data)
        
        # Resumo
        print("\nResumo dos dados filtrados:")
        print(filtered_data.describe().transpose())
        print(f"\nProcesso concluído com sucesso!\nArquivo salvo em: {OUTPUT_FILE}")
    
    except Exception as e:
        print(f"\nErro durante a execução: {str(e)}")
        raise

if __name__ == "__main__":
    main()